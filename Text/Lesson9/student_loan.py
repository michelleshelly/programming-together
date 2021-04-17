#define function - code i can repeat
# function - create_student_loan_workbook
# class - Workbook()
# method - create sheet 
# fields - data lives in object (invisible)

from openpyxl import Workbook

def create_student_loan_workbook():
    student_loan = Workbook()
    loan_repayment_plan = student_loan.active
    loan_repayment_plan.title = "Loan Repayment Plan"
    #loan_repayment_plan = student_loan.create_sheet("Loan Repayment Plan")
    headerNames = [
        "Time",
        "Remaining Principal",
        "Payment Per Month",
        "Interest Payable Per Month",
        "Interest Rate (per annum)",
        "Principal Paid"
    ]
    loan_repayment_plan.append(headerNames)
    return student_loan

def remaining_principal_is_positive(row):
    return row[1]>0
    
def computeRow(previousRow:list):
    time = previousRow[0]
    remaining_principal = previousRow[1]
    payment_per_month = previousRow[2]
    interest_payable_that_month = previousRow[3]
    interest_rate = previousRow[4]
    principal_paid = previousRow[5]

    new_remaining_principal = remaining_principal - principal_paid
    new_interest_payable = (remaining_principal - principal_paid) * interest_rate/12

    if payment_per_month > (new_remaining_principal + new_interest_payable):
        payment_per_month= new_remaining_principal + new_interest_payable
    new_principal_paid = payment_per_month - new_interest_payable

    nextRow = [
        time + 1,
        new_remaining_principal,
        payment_per_month,
        new_interest_payable,
        interest_rate,
        new_principal_paid
    ]

    return nextRow

student_loan = create_student_loan_workbook()
loan_repayment_plan = student_loan.get_sheet_by_name("Loan Repayment Plan")
#same as loan_repayment_plan = student.loan.active (the above is if there are multiple sheets)

initial_principal =24000
interest_rate = 4.75/100
payment_per_month = 1500

row = [0, initial_principal, payment_per_month, 0, interest_rate, 0]
while (remaining_principal_is_positive(row)):
    loan_repayment_plan.append(row)
    #to add the computed row
    row = computeRow(row)


#loan_repayment_plan = student_loan.active
    #to create on the first sheet
#loan_repayment_plan.title = "Loan Repayment Plan"
    #title means the name of that sheet

student_loan.save("Text/Lesson9/Student Loan For Michelle.xlsx")