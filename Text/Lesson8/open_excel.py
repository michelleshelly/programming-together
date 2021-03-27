from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['b1'] = "hello"
ws['c1'] = "im"
ws['d1'] = "michelle"

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A1'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")
